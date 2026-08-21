"""#SMETA-7: экспорт Smeta в печатную форму .xlsx («Смета 12 гр. по ФЕР»).

Работает с уже загруженной моделью sv.model.Smeta, а не с исходным файлом — поэтому
одинаково экспортирует .xlsx/.sobx/.arp/ЛСР в единый привычный вид печатной формы.
Раскладка колонок и стиль сверены на печатной форме реальной сметы (файл заказчика,
в репозиторий не входит), 2026-08-21.

ЕДИНИЦЫ: в модели все деньги — РУБЛИ. В шапке итоги печатаются в ТЫСЯЧАХ рублей
(как в оригинальной форме), поэтому шапка делит на 1000 при выводе; тело таблицы —
в рублях, без деления. Не путать эти два места.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from sv.model import Smeta


def _fmt_money(v: float | None) -> str:
    if v is None:
        return ""
    return f"{v:,.2f}".replace(",", " ").replace(".", ",")


def _fmt_qty(v: float | None) -> str:
    if v is None:
        return ""
    if v.is_integer():
        return f"{v:.0f}"
    return f"{v:g}"


def export_smeta_xlsx(smeta: Smeta, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"

    # Ширины колонок
    for col, width in enumerate([5, 13, 55, 9, 9, 11, 9, 13, 11, 9, 13, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    row = 1

    # Шапка
    title = next((v for v in [smeta.smeta_num, smeta.work_name, smeta.object_name] if v), "Смета")
    ws.merge_cells(f"A{row}:L{row}")
    cell = ws[f"A{row}"]
    cell.value = title
    cell.font = Font(bold=True, size=13)
    cell.alignment = Alignment(horizontal="center")
    row += 2

    if smeta.object_name and smeta.object_name != title:
        ws.merge_cells(f"A{row}:L{row}")
        cell = ws[f"A{row}"]
        cell.value = smeta.object_name
        cell.font = Font(italic=True)
        row += 1

    row += 1  # Пустая строка

    # Таблица итогов
    for label, base, current in smeta.totals.rows():
        ws[f"C{row}"] = label
        if base is not None:
            ws[f"H{row}"] = f"{_fmt_money(base / 1000)} тыс.руб."
        else:
            ws[f"H{row}"] = ""
        if current is not None:
            ws[f"K{row}"] = f"{_fmt_money(current / 1000)} тыс.руб."
        else:
            ws[f"K{row}"] = ""
        if "Сметная стоимость" in label:
            ws[f"C{row}"].font = Font(bold=True)
        row += 1

    row += 1  # Пустая строка

    # Заголовок таблицы позиций
    header = [
        "№", "Шифр расценки", "Наименование работ и затрат", "Ед.изм.", "Кол-во",
        "Цена на ед.изм.", "Коэфф.", "Стоимость баз.цены", "Пункт коэфф.пересчёта",
        "Коэфф.пересчёта", "Стоимость в текущих ценах", "ЗТР всего чел.-ч"
    ]
    for i, value in enumerate(header, 1):
        ws.cell(row=row, column=i, value=value)
        ws.cell(row=row, column=i).font = Font(bold=True)
        ws.cell(row=row, column=i).alignment = Alignment(wrap_text=True)
    row += 1

    # Нумерация колонок
    for i in range(1, 13):
        cell = ws.cell(row=row, column=i)
        cell.value = str(i)
        cell.font = Font(italic=True, size=8)
    row += 1

    ws.freeze_panes = f"A{row}"

    # Тело таблицы
    for section in smeta.sections():
        ws.merge_cells(f"A{row}:L{row}")
        cell = ws[f"A{row}"]
        cell.value = f"Раздел: {section.name}"
        cell.font = Font(bold=True, size=11)
        row += 1

        prev_subsection = None
        for position in section.positions:
            if position.subsection and position.subsection != prev_subsection:
                ws.merge_cells(f"A{row}:L{row}")
                cell = ws[f"A{row}"]
                cell.value = f"Подраздел: {position.subsection}"
                cell.font = Font(bold=True, italic=True, size=10)
                row += 1
                prev_subsection = position.subsection

            # Строка позиции
            ws[f"A{row}"] = position.num
            ws[f"B{row}"] = position.code
            ws[f"C{row}"] = position.name
            ws[f"C{row}"].alignment = Alignment(wrap_text=True)
            ws[f"D{row}"] = position.unit
            ws[f"E{row}"] = _fmt_qty(position.qty)
            ws[f"F{row}"] = _fmt_money(position.price_base)
            row += 1

            # Ресурсные строки
            for label, base, current in position.resources.rows():
                ws[f"C{row}"] = f"    {label}"
                ws[f"C{row}"].font = Font(italic=True, color="999999")
                ws[f"H{row}"] = _fmt_money(base)
                ws[f"K{row}"] = _fmt_money(current)
                row += 1

            # Примечание
            if position.note:
                ws[f"C{row}"] = position.note
                ws[f"C{row}"].font = Font(italic=True, size=8)
                row += 1

            # Итог позиции
            ws[f"H{row}"] = _fmt_money(position.total_base)
            ws[f"K{row}"] = _fmt_money(position.total_current)
            ws[f"L{row}"] = _fmt_qty(position.labor)
            for col in ["H", "K", "L"]:
                ws[f"{col}{row}"].font = Font(bold=True)
            row += 1

    # Печать
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.5
    ws.page_margins.top = ws.page_margins.bottom = 0.75

    wb.save(path)
