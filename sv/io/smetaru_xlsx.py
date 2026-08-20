"""#SMETA-1: загрузчик формы «Смета 12 гр. по ФЕР» (экспорт .xlsx из Смета.РУ).

Раскладка колонок и правила разбора проверены на реальных файлах формы.
и проверены на реальном файле формы.

ЕДИНИЦЫ: в шапке (строки 1..31) итоги напечатаны в ТЫСЯЧАХ рублей, в позициях — в
рублях. Загрузчик приводит шапку к рублям (умножает на 1000), чтобы модель была
одноединичной. Это самая частая ошибка при разборе таких файлов: смешение даёт ошибку в 1000 раз.
"""
from __future__ import annotations

from openpyxl import load_workbook

from sv.model import Position, Resources, Smeta, SmetaFormat, Totals


def _f(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return None


def _s(value) -> str:
    return "" if value is None else str(value).strip()


def load(path: str) -> Smeta:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    wb.close()

    def cell(r: int, c: int) -> object:
        if 1 <= r <= len(rows) and 1 <= c <= len(rows[r - 1]):
            return rows[r - 1][c - 1]
        return None

    # Шапка
    object_name = None
    work_name = None
    smeta_num = None
    total_base = None
    total_current = None
    construction_base = None
    construction_current = None
    installation_base = None
    installation_current = None
    equipment_base = None
    equipment_current = None
    other_base = None
    other_current = None
    labor_base = None
    labor_current = None
    wages_base = None
    wages_current = None

    for r in range(1, 32):
        a = _s(cell(r, 1))
        b = _s(cell(r, 2))
        c = _s(cell(r, 3))

        # Подпись поля и его ЗНАЧЕНИЕ лежат в РАЗНЫХ строках: значение — строкой
        # выше, в колонке A; сама подпись «(наименование стройки)» — в колонке B.
        # Чтение значения из той же строки давало пустоту всегда, и вкладка
        # называлась именем файла вместо номера сметы.
        if "наименование стройки" in b:
            object_name = _s(cell(r - 1, 1))
        elif "наименование работ и затрат" in b:
            work_name = _s(cell(r - 1, 1))
        elif "ЛОКАЛЬНАЯ СМЕТА" in b:      # номер сметы — в колонке B, не в C
            smeta_num = b
        elif "Сметная стоимость" in c:
            total_base = _f(cell(r, 7))
            total_current = _f(cell(r, 9))
        elif "Строительные работы" in c:
            construction_base = _f(cell(r, 7))
            construction_current = _f(cell(r, 9))
        elif "Монтажные работы" in c:
            installation_base = _f(cell(r, 7))
            installation_current = _f(cell(r, 9))
        elif "Оборудование" in c:
            equipment_base = _f(cell(r, 7))
            equipment_current = _f(cell(r, 9))
        elif "Прочие работы" in c:
            other_base = _f(cell(r, 7))
            other_current = _f(cell(r, 9))
        elif "Нормативная трудоемкость" in c:
            labor_base = _f(cell(r, 7))
            labor_current = _f(cell(r, 9))
        elif "Средства на оплату труда" in c:
            wages_base = _f(cell(r, 7))
            wages_current = _f(cell(r, 9))

    # Приведение шапки к рублям
    def mul1000(v):
        return v * 1000 if v is not None else None

    total_base = mul1000(total_base)
    total_current = mul1000(total_current)
    construction_base = mul1000(construction_base)
    construction_current = mul1000(construction_current)
    installation_base = mul1000(installation_base)
    installation_current = mul1000(installation_current)
    equipment_base = mul1000(equipment_base)
    equipment_current = mul1000(equipment_current)
    other_base = mul1000(other_base)
    other_current = mul1000(other_current)
    wages_base = mul1000(wages_base)
    wages_current = mul1000(wages_current)

    # Позиции
    positions = []
    current_section = ""      # не None: поле модели объявлено str, а None
    current_subsection = ""   # ронял match_key() и прятал позиции из дерева
    last = None
    skipped = 0
    section_totals: dict[str, tuple[float | None, float | None]] = {}

    for r in range(32, len(rows) + 1):
        row = rows[r - 1]
        a = _s(row[0])
        b = _s(row[1])
        c = _s(row[2])
        d = _s(row[3])
        e = _f(row[4])
        f = _f(row[5])
        g = _s(row[6])
        h = _f(row[7])
        i = _s(row[8])
        j = _f(row[9])
        k = _f(row[10])
        l = _f(row[11])

        if a == "" and b == "" and c == "":
            # ВНИМАНИЕ: колонка G двузначна. В строке ПОЗИЦИИ это коэффициент-выражение
            # (")*1,22"), поэтому выше она читается как строка; в строке ИТОГА той же
            # формы — базисная стоимость, число. Читать её здесь через _s() значит
            # положить в модель строку: sum_positions_base() падал с TypeError.
            g_num = _f(row[6])
            if last is not None and g_num is not None and j is not None:
                last.total_base = g_num
                last.total_current = j
                last.labor = l
            # Полностью пустая строка — разделитель формы, а не потеря данных:
            # в счётчик нераспознанных не идёт, иначе предупреждение зашумлено.
            continue

        # Название раздела нормализуем: печатная форма пишет «Раздел: Вентиляция»,
        # .sobx и .arp — просто «Вентиляция». Без единой формы сравнение двух
        # редакций одной сметы не находит НИ ОДНОГО совпадения (#SMETA-5).
        if a.startswith("Раздел:"):
            current_section = a.split(":", 1)[1].strip()
            current_subsection = ""   # иначе разделы без своих подразделов
            continue                  # наследуют подраздел предыдущего раздела

        if a.startswith("Подраздел:"):
            current_subsection = a.split(":", 1)[1].strip()
            continue

        # «Итого по разделу: …» / «Итого по подразделу: …» — напечатанные итоги формы.
        # По ним UI показывает суммы разделов и сверяет их с суммой позиций.
        if a.startswith("Итого по"):
            section_totals[a] = (_f(row[6]), j)   # G здесь — число, не коэффициент
            continue

        if a.startswith("Локальная смета"):
            continue

        if a.isdigit():
            pos = Position(
                num=a,
                code=b,
                name=c,
                unit=d,
                qty=e,
                price_base=f,
                coef=g or None,
                cost_base=h,
                index_point=i or None,
                index=j,
                cost_current=k,
                section=current_section,
                subsection=current_subsection,
                source_row=r
            )
            positions.append(pos)
            last = pos
            continue

        if last is not None:
            # Ресурсная строка
            if c == "Зарплата":
                last.resources = Resources(
                    zarplata_base=h,
                    zarplata_current=k
                )
                if g:
                    last.coef = g
            elif c == "Эксплуатация машин":
                if not last.resources:
                    last.resources = Resources()
                last.resources.ekspl_mashin_base = h
                last.resources.ekspl_mashin_current = k
            elif c == "в т.ч. зарплата машинистов":
                if not last.resources:
                    last.resources = Resources()
                last.resources.zp_mashinistov_base = h
                last.resources.zp_mashinistov_current = k
            elif c == "Материальные ресурсы":
                if not last.resources:
                    last.resources = Resources()
                last.resources.materialy_base = h
                last.resources.materialy_current = k
            elif c == "Затраты труда":
                if not last.resources:
                    last.resources = Resources()
                last.resources.zatraty_truda_qty = e
                last.resources.zatraty_truda_value = l
                if g:
                    last.coef = g
            elif c.startswith("НР от"):
                if not last.resources:
                    last.resources = Resources()
                last.resources.nr_percent = e
                last.resources.nr_base = h
                last.resources.nr_current = k
            elif c.startswith("Объем:"):
                # пояснение формы: расценка на 10/100 единиц при дробном объёме
                last.note = c
            elif c.startswith("СП от"):
                if not last.resources:
                    last.resources = Resources()
                last.resources.sp_percent = e
                last.resources.sp_base = h
                last.resources.sp_current = k
            else:
                skipped += 1
        else:
            skipped += 1

    totals = Totals(
        total_base=total_base,
        total_current=total_current,
        construction_base=construction_base,
        construction_current=construction_current,
        installation_base=installation_base,
        installation_current=installation_current,
        equipment_base=equipment_base,
        equipment_current=equipment_current,
        other_base=other_base,
        other_current=other_current,
        labor_base=labor_base,
        labor_current=labor_current,
        wages_base=wages_base,
        wages_current=wages_current
    )

    smeta = Smeta(
        path=path,
        fmt=SmetaFormat.SMETARU_XLSX,
        object_name=object_name,
        work_name=work_name,
        smeta_num=smeta_num,
        positions=positions,
        totals=totals
    )

    smeta.section_totals = section_totals
    if skipped > 0:
        smeta.warnings.append(f"не распознано строк: {skipped}")

    if not positions:
        smeta.warnings.append("не найдено ни одной позиции — возможно, это не форма «Смета 12 гр. по ФЕР»")

    return smeta
