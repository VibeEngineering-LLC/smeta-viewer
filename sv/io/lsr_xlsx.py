"""#SMETA-6: загрузчик входящей ЛСР по Методике 421/пр (ГРАНД-Смета/РИМ, .xlsx).

Раскладка: шапка строки 1..38, тело с 39. A=№, B=обоснование, C=наименование,
H=ед.изм., I=кол-во, J=коэффициент, K=кол-во с учётом коэффициента, L=цена
базисная за ед., M=индекс пересчёта, N=цена текущая за ед., O=всего текущая
(итог позиции — берётся из строки "Всего по позиции", НЕ из строки самой
позиции). Под позицией — строки ФОТ / НР ... / СП ....

ВАЖНО (по опыту у составителя смет и по разбору соседних форматов): итог шапки
"Сметная стоимость" у входящей ЛСР включает лимитированные затраты и НДС, а
сумма по позициям — НЕТ. Расхождение 15-25% между ними для ЭТОГО формата —
норма, а не ошибка; сверивать их как "должны совпадать" нельзя. Отдельно от
итога — единица шапки не считается тысячами по умолчанию: она читается из
файла (текст рядом со значением), потому что смешение единиц уже было
источником ошибки в соседнем формате (Смета.РУ).

НЕ ПРОВЕРЕНО НА РЕАЛЬНОМ ФАЙЛЕ С ИЗВЕСТНЫМИ ИТОГАМИ (в отличие от трёх других
загрузчиков проекта) — только структурно, на синтетическом файле по этой же
раскладке. При первом реальном файле — сверить числа и снять это предупреждение.
"""
from __future__ import annotations

from openpyxl import load_workbook

from sv.model import Position, Resources, Smeta, SmetaFormat, Totals


def _f(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        s = str(value).strip().replace(",", ".")
        return float(s)
    except (ValueError, TypeError):
        return None


def _s(value) -> str:
    return "" if value is None else str(value).strip()


def _is_thousands(unit_text: str) -> bool:
    return "тыс" in unit_text.lower()


def looks_like_lsr(path: str) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        for r in range(1, 16):
            a = _s(ws.cell(row=r, column=1).value)
            b = _s(ws.cell(row=r, column=2).value)
            if "Наименование программного продукта" in (a or "") or "Наименование программного продукта" in (b or ""):
                wb.close()
                return True
        wb.close()
        return False
    except Exception:
        return False


def load(path: str) -> Smeta:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    wb.close()

    def cell(r: int, c: int):
        if 1 <= r <= len(rows) and 1 <= c <= len(rows[r - 1]):
            return rows[r - 1][c - 1]
        return None

    program = ""
    region = ""
    object_name = ""
    method = ""
    total_cost = total_construction = total_installation = total_equipment = total_other = None

    for r in range(1, 39):
        a = _s(cell(r, 1))
        b = _s(cell(r, 2))

        if "Наименование программного продукта" in (a or "") or "Наименование программного продукта" in (b or ""):
            for col in range(3, 6):
                val = _s(cell(r, col))
                if val:
                    program = val
                    break

        elif "Наименование субъекта Российской Федерации" in (a or ""):
            for col in range(3, 6):
                val = _s(cell(r, col))
                if val:
                    region = val
                    break

        elif "(наименование работ и затрат)" in (a or ""):
            object_name = _s(cell(r - 1, 1))

        elif "Составлен" in (a or "") and "методом" in (" ".join([_s(cell(r, col)) for col in range(1, 5)]) or ""):
            method = " ".join([_s(cell(r, col)) for col in range(1, 5) if _s(cell(r, col))])

        elif "Сметная стоимость" in (a or "") or "Сметная стоимость" in (b or ""):
            raw = _f(cell(r, 4))
            unit_text = _s(cell(r, 5))
            total_cost = raw * 1000 if raw is not None and _is_thousands(unit_text) else raw

        elif "строительных работ" in (a or "") or "строительных работ" in (b or ""):
            raw = _f(cell(r, 4))
            unit_text = _s(cell(r, 5))
            total_construction = raw * 1000 if raw is not None and _is_thousands(unit_text) else raw

        elif "монтажных работ" in (a or "") or "монтажных работ" in (b or ""):
            raw = _f(cell(r, 4))
            unit_text = _s(cell(r, 5))
            total_installation = raw * 1000 if raw is not None and _is_thousands(unit_text) else raw

        elif "оборудования" in (a or "") or "оборудования" in (b or ""):
            raw = _f(cell(r, 4))
            unit_text = _s(cell(r, 5))
            total_equipment = raw * 1000 if raw is not None and _is_thousands(unit_text) else raw

        elif "прочих затрат" in (a or "") or "прочих затрат" in (b or ""):
            raw = _f(cell(r, 4))
            unit_text = _s(cell(r, 5))
            total_other = raw * 1000 if raw is not None and _is_thousands(unit_text) else raw

    positions = []
    current_section = ""
    last: Position | None = None

    for r in range(39, len(rows) + 1):
        a = _s(cell(r, 1))
        b = _s(cell(r, 2))
        c = _s(cell(r, 3))
        h = _s(cell(r, 8))
        i = _f(cell(r, 9))
        j = _s(cell(r, 10))
        k = _f(cell(r, 11))
        l = _f(cell(r, 12))
        m = _f(cell(r, 13))
        n = _f(cell(r, 14))
        o = _f(cell(r, 15))

        if a.startswith("Раздел"):
            current_section = a
            continue

        if a.isdigit() and b:
            pos = Position(
                num=a,
                code=b,
                name=c,
                unit=h,
                qty=k,
                price_base=l,
                index=m,
                section=current_section,
                source_row=r,
                coef=(j or None),
            )
            positions.append(pos)
            last = pos
        elif c.strip() == "Всего по позиции" and last is not None:
            last.total_current = o
        elif a == "" and c and last is not None:
            if c == "ФОТ":
                last.resources.zarplata_current = o
            elif c.startswith("НР "):
                last.resources.nr_current = o
            elif c.startswith("СП "):
                last.resources.sp_current = o

    totals = Totals(
        total_current=total_cost,
        construction_current=total_construction,
        installation_current=total_installation,
        equipment_current=total_equipment,
        other_current=total_other,
    )

    warnings = [
        "итог шапки ЛСР включает лимитированные затраты и НДС — сравнивать с суммой позиций напрямую нельзя (расхождение 15-25% для этого формата норма)"
    ]

    if not positions:
        warnings.append("не найдено позиций — возможно, это не форма ЛСР по Методике 421/пр")

    return Smeta(
        path=path,
        fmt=SmetaFormat.LSR_XLSX,
        object_name=object_name,
        work_name=method,
        smeta_num=region,
        positions=positions,
        totals=totals,
        warnings=warnings,
    )
