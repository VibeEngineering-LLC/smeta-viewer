"""Тесты экспорта сметы в Excel/HTML/PDF (#SMETA-7).

Смета — синтетическая фикстура, собранная прямо в тесте: два раздела, у первого
позиция БЕЗ подраздела, у обоих разделов — позиция с ПОДРАЗДЕЛОМ ОДНОГО И ТОГО ЖЕ
имени. Оба случая — пойманные при разработке дефекты (ложная строка «Подраздел: »
для позиции без подраздела; несброшенный между разделами счётчик «предыдущий
подраздел», из-за которого одноимённый подраздел второго раздела не печатался).
"""
from __future__ import annotations

import os

import pytest
from openpyxl import load_workbook

from sv.model import Position, Resources, Section, Smeta, SmetaFormat, Totals
from sv.io.export_xlsx import export_smeta_xlsx
from sv.ui.print_export import build_html, export_pdf


@pytest.fixture
def smeta() -> Smeta:
    """Синтетическая смета для тестов."""
    pos1 = Position(
        num="1",
        code="20-02-016-01",
        name="Установка пластин",
        unit="ШТ",
        qty=8,
        price_base=12.78,
        total_base=218.97,
        total_current=12322.38,
        labor=6.72,
        section="Раздел А",
        subsection="",
        resources=Resources(
            zarplata_base=59.52,
            zarplata_current=3989.03,
            nr_percent=121,
            nr_base=73.18,
            nr_current=4904.58
        )
    )
    pos2 = Position(
        num="2",
        code="20-06-005-01",
        name="Установка фильтров",
        unit="ШТ",
        qty=3,
        price_base=131.06,
        total_base=915.90,
        total_current=55263.80,
        labor=27.96,
        section="Раздел А",
        subsection="Общие работы",
        resources=Resources(
            materialy_base=86.37,
            materialy_current=21963.56
        )
    )
    pos3 = Position(
        num="3",
        code="20-01-002-05",
        name="Прокладка воздуховодов",
        unit="100 м2",
        qty=0.3274,
        price_base=3037.22,
        total_base=100.0,
        total_current=110.0,
        labor=1.0,
        section="Раздел Б",
        subsection="Общие работы"
    )
    # Позиция БЕЗ подраздела, идущая ПОСЛЕ подраздельной. Именно этот порядок ловит
    # ложный заголовок «Подраздел: » с пустым именем: когда позиция без подраздела
    # стоит первой в разделе, дефект не проявляется — начальное значение «предыдущего
    # подраздела» совпадает с пустым, и мутационная проверка проходила зелёной при
    # заведомо сломанном условии.
    pos4 = Position(
        num="4",
        code="20-01-002-06",
        name="Позиция без подраздела",
        unit="шт",
        qty=1,
        price_base=10.0,
        total_base=10.0,
        total_current=11.0,
        labor=0.5,
        section="Раздел Б",
        subsection=""
    )
    return Smeta(
        path="synthetic.xlsx",
        fmt=SmetaFormat.SMETARU_XLSX,
        object_name="Тестовый объект",
        work_name="Комплекс работ",
        smeta_num="Локальная смета №1",
        positions=[pos1, pos2, pos3, pos4],
        totals=Totals(
            total_base=1234.87,
            total_current=68500.18,
            construction_base=1234.87,
            construction_current=68500.18
        )
    )


def test_export_xlsx_creates_valid_file(smeta, tmp_path):
    """Тест экспорта в Excel: файл создаётся и содержит нужные данные."""
    path = str(tmp_path / "out.xlsx")
    export_smeta_xlsx(smeta, path)
    wb = load_workbook(path)
    ws = wb.active
    values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert "20-02-016-01" in values
    assert "20-06-005-01" in values
    assert "20-01-002-05" in values
    assert any("Локальная смета №1" in val for val in values)


def test_export_xlsx_no_empty_subsection_header(smeta, tmp_path):
    """Регрессия: не должно быть пустого заголовка подраздела."""
    path = str(tmp_path / "out.xlsx")
    export_smeta_xlsx(smeta, path)
    wb = load_workbook(path)
    ws = wb.active
    values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert "Подраздел: " not in values


def test_export_xlsx_subsection_printed_twice(smeta, tmp_path):
    """Регрессия: подраздел должен печататься по разу в каждом разделе."""
    path = str(tmp_path / "out.xlsx")
    export_smeta_xlsx(smeta, path)
    wb = load_workbook(path)
    ws = wb.active
    values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    count = values.count("Подраздел: Общие работы")
    assert count == 2


def test_build_html_contains_positions(smeta):
    """Тест HTML-экспорта: содержит коды позиций и названия разделов."""
    html = build_html(smeta)
    assert "20-02-016-01" in html
    assert "20-06-005-01" in html
    assert "20-01-002-05" in html
    assert "Раздел А" in html
    assert "Раздел Б" in html


def test_build_html_no_empty_subsection_header(smeta):
    """Регрессия: не должно быть пустого заголовка подраздела в HTML."""
    html = build_html(smeta)
    assert "Подраздел: </td>" not in html


def test_build_html_subsection_printed_twice(smeta):
    """Регрессия: подраздел должен печататься по разу в каждом разделе в HTML."""
    html = build_html(smeta)
    assert html.count("Подраздел: Общие работы") == 2


def test_build_html_escapes_html_special_chars():
    """Тест экранирования специальных символов HTML."""
    pos = Position(
        num="1",
        code="",
        name='Клапан <A&B> "тест"',
        unit="ШТ",
        qty=0,
        price_base=0,
        total_base=0,
        total_current=0,
        labor=0,
        section="Раздел",
        subsection=""
    )
    smeta = Smeta(
        path="",
        fmt=SmetaFormat.SMETARU_XLSX,
        object_name="",
        work_name="",
        smeta_num="",
        positions=[pos],
        totals=Totals(total_base=0, total_current=0, construction_base=0, construction_current=0)
    )
    html = build_html(smeta)
    assert "<A&B>" not in html
    assert "&lt;A&amp;B&gt;" in html


def test_export_pdf_creates_nonempty_file(smeta, tmp_path):
    """Тест экспорта в PDF: файл создаётся и не пустой."""
    pytest.importorskip("PySide6.QtWidgets")
    from PySide6 import QtWidgets
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([])
    path = str(tmp_path / "out.pdf")
    export_pdf(smeta, path)
    assert os.path.exists(path)
    assert os.path.getsize(path) > 0
